from streamlit.testing.v1 import AppTest

from openpyxl import load_workbook

from vendor_agent.uploads import UploadedArtifact, stage_uploaded_case

from app import _markdown_brief, _matching_sample_baseline, _resolved_missing_items, _triage_workbook_bytes
from vendor_agent.pipeline import run_case


PACKAGE_ROOT = "data/source-package/Candidate_package"


def test_streamlit_app_renders_dashboard_homepage():
    app = AppTest.from_file("app.py")
    app.run(timeout=30)

    assert not app.exception
    assert any("Vendor Onboarding Triage" in item.value for item in app.title)
    assert any("Vendor Requests" in item.value for item in app.subheader)
    assert any("Queue Priorities" in item.value for item in app.subheader)
    assert any("Missing Items" in item.label for item in app.metric)
    assert len(buttons_by_label(app, "Open")) == 3


def test_streamlit_app_opens_seeded_request_from_dashboard():
    app = AppTest.from_file("app.py")
    app.run(timeout=30)
    buttons_by_label(app, "Open")[1].click()
    app.run(timeout=30)

    assert not app.exception
    assert any("Workspace Depot" in item.value for item in app.markdown)
    assert any("Low" in item.value for item in app.metric)
    assert any("Decision" in item.value for item in app.subheader)
    assert any("Required Vendor Follow-up" in item.value for item in app.subheader)
    assert any("Internal Review Route" in item.value for item in app.subheader)
    assert any("Reviewer Brief" in item.value for item in app.subheader)
    assert any("Drafts" in item.label for item in app.expander)
    assert any("Audit details" in item.label for item in app.expander)
    assert any("Delete request" in item.label for item in app.button)
    assert not any("Updated tax form" in checkbox.label for checkbox in app.checkbox)


def test_streamlit_app_submit_request_view_explains_queue_creation():
    app = AppTest.from_file("app.py")
    app.run(timeout=30)
    button_by_label(app, "Submit New Request").click()
    app.run(timeout=30)

    assert not app.exception
    assert any("Submit New Request" in item.value for item in app.subheader)
    assert any("add a new request to the queue" in item.value for item in app.info)
    assert any("Required: intake workbook" in item.value for item in app.caption)


def test_streamlit_app_can_delete_request_from_detail_view():
    app = AppTest.from_file("app.py")
    app.run(timeout=30)
    buttons_by_label(app, "Open")[1].click()
    app.run(timeout=30)
    button_by_label(app, "Delete request").click()
    app.run(timeout=30)

    assert not app.exception
    assert any(item.label == "Requests" and item.value == "2" for item in app.metric)
    assert len(buttons_by_label(app, "Open")) == 2


def test_uploaded_support_artifacts_can_be_compared_to_sample_baseline(tmp_path):
    from pathlib import Path

    baseline = run_case(Path(PACKAGE_ROOT) / "cases" / "case_003")
    uploaded_case = stage_uploaded_case(
        [
            uploaded_fixture("intake.xlsx", "case_003_intake.xlsx"),
            uploaded_fixture("quote.csv", "case_003_quote.csv"),
            uploaded_fixture("contract.pdf", "case_003_contract.pdf"),
            uploaded_fixture("security_questionnaire.md", "case_003_security_questionnaire.md"),
            uploaded_fixture("vendor_email.txt", "case_003_vendor_email.txt"),
            text_artifact("signed-dpa.md", "Executed Data Processing Agreement."),
            text_artifact("soc2-type-ii.md", "Current SOC 2 Type II report."),
            text_artifact(
                "ai-training-opt-out.txt",
                "Confirmed model training opt-out and service improvement disablement.",
            ),
        ],
        tmp_path,
        template_package_root=Path(PACKAGE_ROOT),
    )
    uploaded_packet = run_case(uploaded_case.case_dir)
    baseline_id, _ = _matching_sample_baseline(uploaded_packet)

    assert uploaded_case.is_ready
    assert baseline_id == "case_003"
    assert "Data Processing Agreement" in _resolved_missing_items(baseline, uploaded_packet)
    assert (
        "SOC 2 Type II report or equivalent security attestation"
        in _resolved_missing_items(baseline, uploaded_packet)
    )


def test_uploaded_net_new_packet_has_no_sample_baseline(tmp_path):
    from pathlib import Path

    zip_path = Path("data/sample-upload-packets/zips/net_new_supportflow_complete.zip")
    uploaded_case = stage_uploaded_case(
        [UploadedArtifact(name=zip_path.name, content=zip_path.read_bytes())],
        tmp_path,
        template_package_root=Path(PACKAGE_ROOT),
    )
    uploaded_packet = run_case(uploaded_case.case_dir)
    baseline_id, baseline_packet = _matching_sample_baseline(uploaded_packet)

    assert uploaded_case.is_ready
    assert uploaded_packet.facts.vendor_name == "SupportFlow Assist"
    assert baseline_id is None
    assert baseline_packet is None


def test_markdown_brief_contains_review_packet_sections():
    from pathlib import Path

    packet = run_case(Path("data/source-package/Candidate_package/cases/case_003"))
    brief = _markdown_brief(packet)

    assert "# TalentPulse AI Vendor Triage Brief" in brief
    assert "## Reviewer Brief" in brief
    assert "## Missing Information" in brief
    assert "## Required Human Route" in brief
    assert "## Findings" in brief


def test_triage_workbook_export_contains_packet_sheets():
    from io import BytesIO
    from pathlib import Path

    packet = run_case(Path("data/source-package/Candidate_package/cases/case_003"))
    workbook_bytes = _triage_workbook_bytes(packet)
    wb = load_workbook(BytesIO(workbook_bytes), read_only=True)

    assert {"Summary", "Missing Info", "Findings", "Approval Route", "Trace", "Synthesis"} <= set(wb.sheetnames)
    assert wb["Summary"]["B2"].value == "TalentPulse AI"
    assert wb["Synthesis"]["B3"].value == "passed"


def uploaded_fixture(uploaded_name: str, fixture_name: str) -> UploadedArtifact:
    from pathlib import Path

    return UploadedArtifact(
        name=uploaded_name,
        content=(Path(PACKAGE_ROOT) / "cases" / "case_003" / fixture_name).read_bytes(),
    )


def text_artifact(uploaded_name: str, text: str) -> UploadedArtifact:
    return UploadedArtifact(name=uploaded_name, content=text.encode("utf-8"))


def button_by_label(app: AppTest, label: str):
    for button in app.button:
        if button.label == label:
            return button
    raise AssertionError("Could not find button %r. Buttons: %s" % (label, [button.label for button in app.button]))


def buttons_by_label(app: AppTest, label: str):
    return [button for button in app.button if button.label == label]
