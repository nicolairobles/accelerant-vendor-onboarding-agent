"""Deterministic parsers for the candidate package files."""

import csv
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from pypdf import PdfReader

from .schemas import (
    ContractTerms,
    DocumentChecklistItem,
    ExtractedValue,
    QuoteLineItem,
    QuoteSummary,
    SecurityFacts,
    SourceEvidence,
)


class EvidenceStore:
    def __init__(self) -> None:
        self._items: List[SourceEvidence] = []
        self._counter = 1

    @property
    def items(self) -> List[SourceEvidence]:
        return self._items

    def add(
        self,
        source_file: Path,
        source_type: str,
        location: str,
        snippet: Any,
        parsed_value: Any = None,
        policy_reference: str = None,
    ) -> str:
        evidence = SourceEvidence(
            id="E%04d" % self._counter,
            source_file=str(source_file),
            source_type=source_type,
            location=location,
            snippet=_clean_snippet(snippet),
            parsed_value=None if parsed_value is None else str(parsed_value),
            policy_reference=policy_reference,
        )
        self._counter += 1
        self._items.append(evidence)
        return evidence.id


def parse_intake_workbook(
    path: Path, evidence: EvidenceStore
) -> Tuple[Dict[str, ExtractedValue], List[DocumentChecklistItem]]:
    wb = load_workbook(path, data_only=True)
    values: Dict[str, ExtractedValue] = {}
    checklist: List[DocumentChecklistItem] = []

    ws = wb["Intake Form"]
    rows = list(ws.iter_rows())
    header_index = _find_header_row(rows, "Field Key")
    headers = [_cell_value(cell) for cell in rows[header_index]]
    field_key_col = headers.index("Field Key") + 1
    value_col = headers.index("Value") + 1
    data_type_col = headers.index("Data Type") + 1

    for row_number in range(header_index + 2, ws.max_row + 1):
        key = _cell_value(ws.cell(row_number, field_key_col))
        if not key:
            continue
        value_cell = ws.cell(row_number, value_col)
        data_type = _cell_value(ws.cell(row_number, data_type_col))
        value = _normalize_value(value_cell.value, data_type)
        evidence_id = evidence.add(
            path,
            "xlsx",
            "Intake Form!%s" % value_cell.coordinate,
            value_cell.value,
            value,
        )
        values[key] = ExtractedValue(key=key, value=value, evidence_id=evidence_id)

    ws_checklist = wb["Document Checklist"]
    rows = list(ws_checklist.iter_rows())
    header_index = _find_header_row(rows, "Document Key")
    headers = [_cell_value(cell) for cell in rows[header_index]]
    key_col = headers.index("Document Key") + 1
    provided_col = headers.index("Provided?") + 1
    artifact_col = headers.index("Expected Artifact") + 1
    note_col = headers.index("Candidate Note") + 1

    for row_number in range(header_index + 2, ws_checklist.max_row + 1):
        document_key = _cell_value(ws_checklist.cell(row_number, key_col))
        if not document_key:
            continue
        provided_cell = ws_checklist.cell(row_number, provided_col)
        provided = bool(provided_cell.value)
        artifact = _cell_value(ws_checklist.cell(row_number, artifact_col))
        note = _cell_value(ws_checklist.cell(row_number, note_col))
        evidence_id = evidence.add(
            path,
            "xlsx",
            "Document Checklist!%s" % provided_cell.coordinate,
            "%s | provided=%s | %s" % (document_key, provided, note),
            provided,
        )
        checklist.append(
            DocumentChecklistItem(
                document_key=document_key,
                provided=provided,
                expected_artifact=artifact,
                note=note,
                evidence_id=evidence_id,
            )
        )

    return values, checklist


def parse_quote_csv(path: Path, evidence: EvidenceStore) -> QuoteSummary:
    line_items: List[QuoteLineItem] = []
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            annual_amount = _to_float(row["annual_amount"])
            one_time_amount = _to_float(row["one_time_amount"])
            evidence_id = evidence.add(
                path,
                "csv",
                "row %s" % row_number,
                row,
                "%s annual=%s one_time=%s"
                % (row["line_item"], annual_amount, one_time_amount),
            )
            line_items.append(
                QuoteLineItem(
                    line_item=row["line_item"],
                    billing_type=row["billing_type"],
                    quantity=_to_float(row["quantity"]),
                    unit_price=_to_float(row["unit_price"]),
                    annual_amount=annual_amount,
                    one_time_amount=one_time_amount,
                    notes=row.get("notes", ""),
                    evidence_id=evidence_id,
                )
            )

    return QuoteSummary(
        annual_contract_value=sum(item.annual_amount for item in line_items),
        one_time_fees=sum(item.one_time_amount for item in line_items),
        line_items=line_items,
    )


def parse_contract_pdf(path: Path, evidence: EvidenceStore) -> ContractTerms:
    pages = _extract_pdf_pages(path)
    full_text = "\n".join(text for _, text in pages)
    evidence_ids = [
        evidence.add(path, "pdf", "page %s" % page_number, text)
        for page_number, text in pages
    ]

    return ContractTerms(
        vendor_name=_first_match(full_text, r"Vendor\s+([A-Za-z0-9 .,&-]+)\s+Effective date"),
        effective_date=_first_match(full_text, r"Effective date\s+([0-9-]+)"),
        initial_term_months=_to_int(_first_match(full_text, r"Initial term\s+([0-9]+) months")),
        annual_fees=_to_float(_first_match(full_text, r"Annual fees\s+\$?([0-9,]+)")),
        payment_terms=_first_match(full_text, r"Payment terms\s+(Net [0-9]+)"),
        limitation_of_liability=_first_match(
            full_text, r"Limitation of liability: ([^.]+)\."
        ),
        auto_renewal=_first_match(full_text, r"Auto-renewal:\s*([^.]+)\."),
        data_use=_first_match(full_text, r"Data use:\s*([^.]+)\."),
        deletion_terms=_first_match(full_text, r"Deletion:\s*([^.]+)\."),
        subprocessors=_first_match(full_text, r"Subprocessors:\s*([^.]+)\."),
        evidence_ids=evidence_ids,
    )


def parse_security_questionnaire(path: Path, evidence: EvidenceStore) -> SecurityFacts:
    text = path.read_text(encoding="utf-8")
    sections = _markdown_sections(text)
    evidence_ids = [
        evidence.add(path, "markdown", "section: %s" % name, body)
        for name, body in sections.items()
    ]
    controls = sections.get("Security controls", "")
    soc2 = sections.get("SOC 2", "")

    return SecurityFacts(
        data_processed=_split_list(sections.get("Data processed", "")),
        integrations=_split_list(sections.get("Integrations", "")),
        encryption_in_transit=_bullet_value(controls, "Encryption in transit"),
        encryption_at_rest=_bullet_value(controls, "Encryption at rest"),
        sso_saml=_bullet_value(controls, "SSO/SAML"),
        scim_provisioning=_bullet_value(controls, "SCIM provisioning"),
        audit_logs=_bullet_value(controls, "Audit logs"),
        soc2_type=soc2.strip(),
        soc2_type2_provided="type ii" in soc2.lower() and "not provided" not in soc2.lower(),
        incident_response=sections.get("Incident response", "").strip(),
        subprocessors=_split_list(sections.get("Subprocessors", "")),
        data_retention=sections.get("Data retention", "").strip(),
        ai_model_training=sections.get("AI/model training", "").strip(),
        gaps=_extract_bullets(sections.get("Gaps", "")),
        evidence_ids=evidence_ids,
    )


def parse_vendor_email(path: Path, evidence: EvidenceStore) -> ExtractedValue:
    text = path.read_text(encoding="utf-8")
    evidence_id = evidence.add(path, "text", "full email", text, "vendor_email")
    return ExtractedValue(key="vendor_email", value=text, evidence_id=evidence_id)


def load_policy_documents(paths: Iterable[Path]) -> Dict[str, str]:
    return {path.stem: path.read_text(encoding="utf-8") for path in paths}


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def _find_header_row(rows: List[Tuple[Any, ...]], header_value: str) -> int:
    for index, row in enumerate(rows):
        values = [_cell_value(cell) for cell in row]
        if header_value in values:
            return index
    raise ValueError("Could not find header row containing %r" % header_value)


def _cell_value(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""
    return str(value).strip()


def _normalize_value(value: Any, data_type: str) -> Any:
    if value is None:
        return None
    if data_type == "list":
        return [item.strip() for item in str(value).splitlines() if item.strip()]
    if data_type == "currency":
        return _to_float(value)
    if data_type == "integer":
        return _to_int(value)
    if data_type == "date" and hasattr(value, "date"):
        return value.date().isoformat()
    return str(value).strip()


def _extract_pdf_pages(path: Path) -> List[Tuple[int, str]]:
    reader = PdfReader(str(path))
    pages = []
    for index, page in enumerate(reader.pages, start=1):
        pages.append((index, page.extract_text() or ""))
    return pages


def _markdown_sections(text: str) -> Dict[str, str]:
    sections: Dict[str, str] = {}
    current_name = None
    current_lines: List[str] = []
    for line in text.splitlines():
        if line.startswith("## "):
            if current_name:
                sections[current_name] = "\n".join(current_lines).strip()
            current_name = line.replace("## ", "", 1).strip()
            current_lines = []
        elif current_name:
            current_lines.append(line)
    if current_name:
        sections[current_name] = "\n".join(current_lines).strip()
    return sections


def _split_list(value: str) -> List[str]:
    if not value:
        return []
    value = value.replace(" and ", ", ")
    return [item.strip().strip(".") for item in value.split(",") if item.strip()]


def _bullet_value(text: str, key: str) -> str:
    pattern = r"-\s*%s:\s*(.+)" % re.escape(key)
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def _extract_bullets(text: str) -> List[str]:
    return [line[1:].strip() for line in text.splitlines() if line.strip().startswith("-")]


def _first_match(text: str, pattern: str) -> str:
    match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
    return " ".join(match.group(1).split()) if match else ""


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(str(value).replace("$", "").replace(",", "").strip())


def _to_int(value: Any) -> int:
    return int(float(str(value).replace(",", "").strip()))


def _clean_snippet(value: Any) -> str:
    if isinstance(value, dict):
        value = ", ".join("%s=%s" % (key, val) for key, val in value.items())
    return " ".join(str(value).split())[:1000]
