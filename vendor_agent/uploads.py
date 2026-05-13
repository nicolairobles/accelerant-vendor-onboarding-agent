"""Upload staging for ad hoc vendor packages."""

import re
import shutil
import zipfile
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook

from .inventory import CASE_SUFFIXES


DEFAULT_PACKAGE_ROOT = Path("data/source-package/Candidate_package")
MAX_UPLOAD_FILES = 40
MAX_TOTAL_UPLOAD_BYTES = 25 * 1024 * 1024

ROLE_MIN_SCORES = {
    "intake": 30,
    "quote": 30,
    "contract": 30,
    "security_questionnaire": 30,
    "vendor_email": 30,
}

ROLE_LABELS = {
    "intake": "intake workbook (.xlsx)",
    "quote": "quote CSV (.csv)",
    "contract": "contract PDF (.pdf)",
    "security_questionnaire": "security questionnaire (.md)",
    "vendor_email": "vendor email (.txt)",
}

OPTIONAL_ARTIFACT_LABELS = {
    "soc2_type2": "SOC 2 Type II report or equivalent",
    "data_processing_agreement": "Data Processing Agreement",
    "subprocessor_list": "Subprocessor list",
    "tax_form": "Tax form",
    "vendor_setup_form": "Vendor setup form",
    "ai_training_opt_out": "AI training opt-out confirmation",
    "incident_response_summary": "Incident response and breach notification summary",
    "statement_of_work": "Statement of work",
}


@dataclass(frozen=True)
class UploadedArtifact:
    name: str
    content: bytes


@dataclass(frozen=True)
class UploadRoleMatch:
    role: str
    uploaded_name: str
    staged_name: str


@dataclass(frozen=True)
class UploadedCase:
    case_dir: Path
    role_matches: List[UploadRoleMatch]
    missing_roles: List[str]
    unmatched_files: List[str]
    warnings: List[str]
    optional_matches: List[UploadRoleMatch] = field(default_factory=list)
    blocking_errors: List[str] = field(default_factory=list)

    @property
    def is_ready(self) -> bool:
        return not self.missing_roles and not self.blocking_errors


def stage_uploaded_case(
    artifacts: Iterable[UploadedArtifact],
    work_root: Path,
    template_package_root: Path = DEFAULT_PACKAGE_ROOT,
    case_id: str = "uploaded_case",
) -> UploadedCase:
    """Create a canonical case folder from uploaded files.

    The downstream pipeline expects stable file names, so upload handling is
    deliberately isolated here. Policy docs and mock internal tools still come
    from the bundled candidate package.
    """

    artifacts_list = list(artifacts)
    initial_errors = _upload_size_errors(artifacts_list)
    expanded = _expand_archives(artifacts_list) if not initial_errors else []
    if expanded and len(expanded) > MAX_UPLOAD_FILES:
        initial_errors.append(
            "Upload package expands to %s files; limit is %s."
            % (len(expanded), MAX_UPLOAD_FILES)
        )

    selected, unmatched_files, warnings = _select_role_matches(expanded)
    optional_selected, optional_unmatched = _select_optional_matches(
        expanded,
        selected_names={artifact.name for artifact in selected.values()},
    )
    optional_names = {artifact.name for artifact in optional_selected.values()}
    unmatched_files = sorted(
        name
        for name in set(unmatched_files + optional_unmatched)
        if name not in optional_names
    )
    blocking_errors = initial_errors + _consistency_errors(selected)
    missing_roles = [role for role in CASE_SUFFIXES if role not in selected]

    case_dir = Path(work_root) / "Candidate_package" / "cases" / case_id
    role_matches: List[UploadRoleMatch] = []
    optional_matches: List[UploadRoleMatch] = []
    if missing_roles or blocking_errors:
        return UploadedCase(
            case_dir=case_dir,
            role_matches=[],
            missing_roles=missing_roles,
            unmatched_files=unmatched_files,
            warnings=warnings,
            optional_matches=[],
            blocking_errors=blocking_errors,
        )

    package_root = case_dir.parent.parent
    _reset_directory(package_root)
    case_dir.mkdir(parents=True, exist_ok=True)
    _copy_supporting_package_files(template_package_root, package_root)

    for role, artifact in selected.items():
        staged_name = "%s_%s" % (case_id, CASE_SUFFIXES[role])
        staged_path = case_dir / staged_name
        staged_path.write_bytes(artifact.content)
        role_matches.append(
            UploadRoleMatch(
                role=role,
                uploaded_name=artifact.name,
                staged_name=staged_name,
            )
        )

    support_dir = case_dir / "supporting_artifacts"
    for optional_key, artifact in optional_selected.items():
        support_dir.mkdir(parents=True, exist_ok=True)
        staged_name = "%s_%s%s" % (
            case_id,
            optional_key,
            _safe_suffix(artifact.name),
        )
        (support_dir / staged_name).write_bytes(artifact.content)
        optional_matches.append(
            UploadRoleMatch(
                role=optional_key,
                uploaded_name=artifact.name,
                staged_name="supporting_artifacts/%s" % staged_name,
            )
        )

    if optional_matches:
        _mark_optional_documents(
            case_dir / ("%s_%s" % (case_id, CASE_SUFFIXES["intake"])),
            optional_matches,
        )

    return UploadedCase(
        case_dir=case_dir,
        role_matches=sorted(role_matches, key=lambda item: item.role),
        missing_roles=[],
        unmatched_files=unmatched_files,
        warnings=warnings,
        optional_matches=sorted(optional_matches, key=lambda item: item.role),
        blocking_errors=[],
    )


def missing_role_labels(roles: Iterable[str]) -> List[str]:
    return [ROLE_LABELS.get(role, role) for role in roles]


def _expand_archives(artifacts: List[UploadedArtifact]) -> List[UploadedArtifact]:
    expanded: List[UploadedArtifact] = []
    for artifact in artifacts:
        if artifact.name.lower().endswith(".zip"):
            expanded.extend(_zip_members(artifact))
        else:
            expanded.append(artifact)
    return expanded


def _zip_members(artifact: UploadedArtifact) -> List[UploadedArtifact]:
    members: List[UploadedArtifact] = []
    with zipfile.ZipFile(BytesIO(artifact.content)) as archive:
        zip_infos = [
            member
            for member in archive.infolist()
            if not member.is_dir() and not member.filename.startswith("__MACOSX/")
        ]
        if len(zip_infos) > MAX_UPLOAD_FILES:
            raise ValueError(
                "Zip contains %s files; limit is %s." % (len(zip_infos), MAX_UPLOAD_FILES)
            )
        expanded_bytes = sum(member.file_size for member in zip_infos)
        if expanded_bytes > MAX_TOTAL_UPLOAD_BYTES:
            raise ValueError(
                "Zip expands to %.1f MB; limit is %.1f MB."
                % (expanded_bytes / 1024 / 1024, MAX_TOTAL_UPLOAD_BYTES / 1024 / 1024)
            )
        for member in zip_infos:
            if member.is_dir() or member.filename.startswith("__MACOSX/"):
                continue
            name = member.filename.replace("\\", "/")
            if name.startswith("/") or ".." in Path(name).parts:
                continue
            members.append(
                UploadedArtifact(
                    name="%s:%s" % (artifact.name, name),
                    content=archive.read(member),
                )
            )
    return members


def _upload_size_errors(artifacts: List[UploadedArtifact]) -> List[str]:
    errors: List[str] = []
    if len(artifacts) > MAX_UPLOAD_FILES:
        errors.append(
            "Upload contains %s files; limit is %s."
            % (len(artifacts), MAX_UPLOAD_FILES)
        )
    total_bytes = sum(len(artifact.content) for artifact in artifacts)
    if total_bytes > MAX_TOTAL_UPLOAD_BYTES:
        errors.append(
            "Upload package is %.1f MB; limit is %.1f MB."
            % (total_bytes / 1024 / 1024, MAX_TOTAL_UPLOAD_BYTES / 1024 / 1024)
        )
    return errors


def _select_role_matches(
    artifacts: List[UploadedArtifact],
) -> Tuple[Dict[str, UploadedArtifact], List[str], List[str]]:
    candidates: Dict[str, List[Tuple[int, UploadedArtifact]]] = {
        role: [] for role in CASE_SUFFIXES
    }
    unmatched: List[str] = []
    warnings: List[str] = []

    for artifact in artifacts:
        scores = {
            role: _score_artifact(artifact, role)
            for role in CASE_SUFFIXES
        }
        best_score = max(scores.values()) if scores else 0
        if best_score <= 0:
            unmatched.append(artifact.name)
            continue
        for role, score in scores.items():
            if score == best_score and score >= ROLE_MIN_SCORES[role]:
                candidates[role].append((score, artifact))
                break
        else:
            unmatched.append(artifact.name)

    selected: Dict[str, UploadedArtifact] = {}
    selected_names = set()
    for role, role_candidates in candidates.items():
        if not role_candidates:
            continue
        ordered = sorted(
            role_candidates,
            key=lambda item: (item[0], _specificity(item[1].name, role)),
            reverse=True,
        )
        selected[role] = ordered[0][1]
        selected_names.add(ordered[0][1].name)
        if len(ordered) > 1:
            warnings.append(
                "Multiple files looked like %s; using %s."
                % (ROLE_LABELS[role], ordered[0][1].name)
            )

    for artifact in artifacts:
        if artifact.name not in selected_names and artifact.name not in unmatched:
            unmatched.append(artifact.name)

    return selected, sorted(unmatched), warnings


def _select_optional_matches(
    artifacts: List[UploadedArtifact],
    selected_names: set,
) -> Tuple[Dict[str, UploadedArtifact], List[str]]:
    optional: Dict[str, UploadedArtifact] = {}
    unmatched: List[str] = []
    for artifact in artifacts:
        if artifact.name in selected_names:
            continue
        key = _optional_artifact_key(artifact)
        if key:
            optional.setdefault(key, artifact)
        else:
            unmatched.append(artifact.name)
    return optional, unmatched


def _score_artifact(artifact: UploadedArtifact, role: str) -> int:
    name = artifact.name.lower()
    suffix = CASE_SUFFIXES[role].rsplit(".", 1)[1]
    if not name.endswith("." + suffix):
        return 0

    text = _text_head(artifact.content)
    score = 10
    role_tokens = {
        "intake": ["intake", "request", "form"],
        "quote": ["quote", "pricing", "order"],
        "contract": ["contract", "agreement", "msa"],
        "security_questionnaire": ["security_questionnaire", "security", "questionnaire"],
        "vendor_email": ["vendor_email", "vendor", "email"],
    }[role]
    score += 8 * sum(1 for token in role_tokens if token in name)

    if role == "quote" and {"line_item", "annual_amount", "one_time_amount"} <= set(_csv_headers(text)):
        score += 40
    elif role == "security_questionnaire":
        if "## security controls" in text or "## soc 2" in text:
            score += 40
    elif role == "vendor_email":
        if "subject:" in text or "from:" in text:
            score += 30
        elif "vendor" in name and "email" in name:
            score += 10
    elif role == "contract":
        if artifact.content.startswith(b"%PDF"):
            score += 20
    elif role == "intake":
        if _workbook_looks_like_intake(artifact.content):
            score += 50
        elif "intake" in name:
            score += 20

    return score


def _workbook_looks_like_intake(content: bytes) -> bool:
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return False
    return "Intake Form" in wb.sheetnames and "Document Checklist" in wb.sheetnames


def _optional_artifact_key(artifact: UploadedArtifact) -> str:
    if "policy" in artifact.name.lower():
        return ""
    text = "%s %s" % (artifact.name.lower(), _text_head(artifact.content))
    normalized = text.replace("-", "_").replace(" ", "_")
    if "soc_2" in normalized or "soc2" in normalized or "type_ii" in normalized:
        return "soc2_type2"
    if "data_processing_agreement" in normalized or re.search(r"\bdpa\b", normalized):
        return "data_processing_agreement"
    if "subprocessor" in normalized:
        return "subprocessor_list"
    if "w_9" in normalized or "tax_form" in normalized or "tax" in normalized:
        return "tax_form"
    if "vendor_setup" in normalized or "setup_form" in normalized:
        return "vendor_setup_form"
    if (
        "ai_training_opt_out" in normalized
        or "training_opt_out" in normalized
        or "model_training_opt_out" in normalized
        or "disable_model_training" in normalized
    ):
        return "ai_training_opt_out"
    if "incident_response" in normalized or "breach_notification" in normalized:
        return "incident_response_summary"
    if "statement_of_work" in normalized or re.search(r"\bsow\b", normalized):
        return "statement_of_work"
    return ""


def _consistency_errors(selected: Dict[str, UploadedArtifact]) -> List[str]:
    prefixes = {
        match.group(1)
        for artifact in selected.values()
        for match in [re.search(r"(case_[0-9]{3})", artifact.name.lower())]
        if match
    }
    if len(prefixes) > 1:
        return [
            "Package appears to mix multiple vendor cases: %s."
            % ", ".join(sorted(prefixes))
        ]
    return []


def _specificity(name: str, role: str) -> int:
    normalized = name.lower().replace("-", "_").replace(" ", "_")
    expected = CASE_SUFFIXES[role].replace(".", "_")
    return len(re.findall(re.escape(role), normalized)) + (1 if expected in normalized else 0)


def _csv_headers(text: str) -> List[str]:
    first_line = text.splitlines()[0] if text.splitlines() else ""
    return [item.strip().lower() for item in first_line.split(",")]


def _text_head(content: bytes) -> str:
    return content[:8192].decode("utf-8", errors="ignore").lower()


def _copy_supporting_package_files(template_package_root: Path, package_root: Path) -> None:
    template_package_root = Path(template_package_root)
    for folder_name in ["docs", "tools"]:
        source = template_package_root / folder_name
        destination = package_root / folder_name
        if not source.exists():
            raise FileNotFoundError("Missing supporting package folder: %s" % source)
        shutil.copytree(source, destination)


def _reset_directory(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def _mark_optional_documents(
    intake_path: Path,
    optional_matches: List[UploadRoleMatch],
) -> None:
    wb = load_workbook(intake_path)
    ws = wb["Document Checklist"]
    rows = list(ws.iter_rows())
    header_index = None
    for index, row in enumerate(rows):
        values = [cell.value for cell in row]
        if "Document Key" in values:
            header_index = index
            break
    if header_index is None:
        wb.save(intake_path)
        return

    headers = [cell.value for cell in rows[header_index]]
    key_col = headers.index("Document Key") + 1
    provided_col = headers.index("Provided?") + 1
    note_col = headers.index("Candidate Note") + 1
    optional_by_key = {match.role: match for match in optional_matches}
    matched_keys = set()

    for row_number in range(header_index + 2, ws.max_row + 1):
        document_key = ws.cell(row_number, key_col).value
        if document_key not in optional_by_key:
            continue
        match = optional_by_key[document_key]
        matched_keys.add(document_key)
        ws.cell(row_number, provided_col).value = True
        ws.cell(row_number, note_col).value = (
            "Provided via uploaded support artifact: %s" % match.uploaded_name
        )

    append_row = ws.max_row + 1
    for document_key, match in sorted(optional_by_key.items()):
        if document_key in matched_keys:
            continue
        ws.cell(append_row, key_col).value = document_key
        ws.cell(append_row, provided_col).value = True
        ws.cell(append_row, note_col).value = (
            "Provided via uploaded support artifact: %s" % match.uploaded_name
        )
        expected_col = None
        for index, value in enumerate(headers, start=1):
            if value == "Expected Artifact":
                expected_col = index
                break
        if expected_col:
            ws.cell(append_row, expected_col).value = OPTIONAL_ARTIFACT_LABELS.get(
                document_key,
                document_key.replace("_", " ").title(),
            )
        append_row += 1

    wb.save(intake_path)


def _safe_suffix(name: str) -> str:
    path = Path(name.split(":", 1)[-1])
    suffix = path.suffix.lower()
    return suffix if suffix and re.match(r"^\.[a-z0-9]+$", suffix) else ".txt"
