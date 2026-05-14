import html
import importlib
import json
import os
import shutil
import tempfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook

import vendor_agent.pipeline as pipeline_module
import vendor_agent.policies as policies_module
import vendor_agent.uploads as uploads_module


# Streamlit Community Cloud can hot-reload app.py while keeping previously
# imported project modules alive. Refresh the internal modules so deployed
# upload behavior always matches the checked-out source revision.
policies_module = importlib.reload(policies_module)
uploads_module = importlib.reload(uploads_module)
pipeline_module = importlib.reload(pipeline_module)

run_case = pipeline_module.run_case
UploadedArtifact = uploads_module.UploadedArtifact
missing_role_labels = uploads_module.missing_role_labels
stage_uploaded_case = uploads_module.stage_uploaded_case


PACKAGE_ROOT = Path("data/source-package/Candidate_package")
CASES_ROOT = Path("data/source-package/Candidate_package/cases")
CASE_OPTIONS = ["case_001", "case_002", "case_003"]
CASE_DISPLAY_NAMES = {
    "case_001": "case_001 - Northstar Analytics",
    "case_002": "case_002 - Workspace Depot",
    "case_003": "case_003 - TalentPulse AI",
}
SECRET_ENV_KEYS = [
    "OPENAI_API_KEY",
    "OPENAI_SYNTHESIS_PROVIDER",
    "OPENAI_SYNTHESIS_MODEL",
]


def _configure_streamlit_secret_env() -> None:
    if os.getenv("OPENAI_API_KEY") and os.getenv("OPENAI_SYNTHESIS_PROVIDER"):
        return
    secrets_paths = [
        Path.home() / ".streamlit" / "secrets.toml",
        Path.cwd() / ".streamlit" / "secrets.toml",
    ]
    if not any(path.exists() for path in secrets_paths):
        return
    try:
        secrets = st.secrets
    except Exception:
        return
    for key in SECRET_ENV_KEYS:
        try:
            value = secrets.get(key)
        except Exception:
            continue
        if value and not os.getenv(key):
            os.environ[key] = str(value)


st.set_page_config(
    page_title="Vendor Onboarding Triage",
    page_icon="",
    layout="wide",
)

_configure_streamlit_secret_env()


def main() -> None:
    _initialize_request_store()

    st.title("Vendor Onboarding Triage")
    st.caption("Procurement request queue and evidence-backed review packets")

    render_sidebar()

    view = st.session_state.get("active_view", "queue")
    active_request_id = st.session_state.get("active_request_id")

    if view == "submit":
        render_submit_request()
    elif view == "detail" and active_request_id in st.session_state.get("vendor_requests", {}):
        render_request_detail(active_request_id)
    else:
        st.session_state["active_view"] = "queue"
        render_dashboard()


@st.cache_data(show_spinner=False)
def _run_sample_case_cached(case_id: str, runtime_signature: str):
    _ = runtime_signature
    return run_case(CASES_ROOT / case_id)


@st.cache_data(show_spinner=False)
def _sample_packets(runtime_signature: str):
    _ = runtime_signature
    return {case_id: run_case(CASES_ROOT / case_id) for case_id in CASE_OPTIONS}


def _initialize_request_store(force: bool = False) -> None:
    runtime_signature = _synthesis_runtime_signature()
    needs_reset = (
        force
        or "vendor_requests" not in st.session_state
        or st.session_state.get("request_store_signature") != runtime_signature
    )
    if not needs_reset:
        return

    packets = _sample_packets(runtime_signature)
    requests = {}
    order = []
    for index, case_id in enumerate(CASE_OPTIONS, start=1):
        packet = packets[case_id]
        requests[case_id] = {
            "request_id": case_id,
            "display_id": "VR-%03d" % index,
            "packet": packet,
            "source": "Seeded request",
            "submitted_at": "Exam seed",
            "uploaded_case": None,
            "workspace": None,
        }
        order.append(case_id)

    st.session_state["vendor_requests"] = requests
    st.session_state["vendor_request_order"] = order
    st.session_state["request_store_signature"] = runtime_signature
    st.session_state["upload_request_counter"] = 0
    st.session_state["next_request_display_number"] = len(CASE_OPTIONS) + 1
    st.session_state["active_view"] = "queue"
    st.session_state["active_request_id"] = None
    st.session_state["upload_feedback"] = None


def render_sidebar() -> None:
    with st.sidebar:
        st.subheader("Workspace")
        if st.button("Vendor Requests", use_container_width=True):
            st.session_state["active_view"] = "queue"
            st.session_state["active_request_id"] = None
            st.rerun()
        if st.button("Submit New Request", type="primary", use_container_width=True):
            st.session_state["active_view"] = "submit"
            st.session_state["active_request_id"] = None
            st.rerun()
        st.caption("%s request(s) in queue" % len(_request_records()))
        if st.session_state.get("active_request_id"):
            record = _request_record(st.session_state["active_request_id"])
            if record:
                st.caption("Open: %s" % record["packet"].facts.vendor_name)
        if st.button("Restore Seeded Requests", use_container_width=True):
            _clear_uploaded_workspaces()
            _initialize_request_store(force=True)
            st.rerun()


def render_dashboard() -> None:
    records = _request_records()
    packets = [record["packet"] for record in records]
    blocked_count = len([packet for packet in packets if packet.status == "blocked"])
    ready_count = len([packet for packet in packets if packet.status != "blocked"])
    missing_count = sum(len(packet.missing_information) for packet in packets)

    st.subheader("Vendor Requests")
    metric_cols = st.columns(4)
    metric_cols[0].metric("Requests", len(records))
    metric_cols[1].metric("Blocked", blocked_count)
    metric_cols[2].metric("Ready For Review", ready_count)
    metric_cols[3].metric("Missing Items", missing_count)
    st.caption("Missing Items is queue-wide across all vendor requests, not an additional request count.")

    if not records:
        st.info("No vendor requests are currently in the queue.")
        if st.button("Restore seeded exam requests", type="primary"):
            _initialize_request_store(force=True)
            st.rerun()
        return

    render_request_queue(records)

    st.subheader("Queue Priorities")
    for packet in sorted(
        packets,
        key=lambda item: (
            item.status != "blocked",
            item.facts.risk.tier != "high",
            -len(item.missing_information),
        ),
    ):
        top_action = _primary_next_action(packet)
        st.write(
            "**%s**: %s. Next: %s."
            % (packet.facts.vendor_name, _status_label(packet), top_action)
        )


def render_request_queue(records: list) -> None:
    header = st.columns([0.75, 2.1, 1.05, 0.8, 0.9, 0.9, 1.3, 0.9, 0.9])
    header[0].markdown("**Request**")
    header[1].markdown("**Vendor**")
    header[2].markdown("**Status**")
    header[3].markdown("**Risk**")
    header[4].markdown("**Budget**")
    header[5].markdown("**Missing**")
    header[6].markdown("**Next owner**")
    header[7].markdown("**Source**")
    header[8].markdown("**Actions**")
    for record in records:
        packet = record["packet"]
        row = st.columns([0.75, 2.1, 1.05, 0.8, 0.9, 0.9, 1.3, 0.9, 0.9])
        row[0].write(record["display_id"])
        row[1].markdown("**%s**" % packet.facts.vendor_name)
        row[1].caption(packet.facts.requesting_team)
        row[2].write(_status_label(packet))
        row[3].write(packet.facts.risk.tier.title())
        row[4].write(packet.facts.budget.status.title())
        row[5].write(len(packet.missing_information))
        row[6].write(_next_owner(packet))
        row[7].write(record["source"])
        if row[8].button(
            "Open %s" % packet.facts.vendor_name,
            key="open_%s" % record["request_id"],
            use_container_width=True,
        ):
            st.session_state["active_view"] = "detail"
            st.session_state["active_request_id"] = record["request_id"]
            st.rerun()


def render_request_detail(request_id: str) -> None:
    record = _request_record(request_id)
    if not record:
        st.session_state["active_view"] = "queue"
        st.session_state["active_request_id"] = None
        st.rerun()

    packet = record["packet"]
    nav_cols = st.columns([1, 1, 5])
    if nav_cols[0].button("Back to requests", use_container_width=True):
        st.session_state["active_view"] = "queue"
        st.session_state["active_request_id"] = None
        st.rerun()
    if nav_cols[1].button("Delete request", use_container_width=True):
        _delete_request(request_id)
        st.session_state["active_view"] = "queue"
        st.session_state["active_request_id"] = None
        st.rerun()
    nav_cols[2].caption(
        "%s • %s • Submitted: %s"
        % (record["display_id"], record["source"], record["submitted_at"])
    )

    render_packet(packet, request_id=request_id)
    if record.get("uploaded_case"):
        with st.expander("Upload intake details", expanded=False):
            render_package_delta(packet, record["uploaded_case"], show_title=False)
        render_upload_details(record["uploaded_case"], expanded=False)


def render_submit_request() -> None:
    st.subheader("Submit New Request")
    st.info("Upload a complete vendor package to add a new request to the queue.")
    uploaded_files = st.file_uploader(
        "New package files",
        type=["xlsx", "csv", "pdf", "md", "txt", "zip"],
        accept_multiple_files=True,
        help=(
            "Upload intake workbook, quote CSV, contract PDF, security "
            "questionnaire, vendor email, and optional support artifacts."
        ),
    )
    st.caption("Required: intake workbook, quote CSV, contract PDF, security questionnaire, vendor email.")
    st.caption("Optional: DPA, SOC 2, subprocessors, tax form, vendor setup form, AI opt-out confirmation.")
    run_clicked = st.button("Add request to queue", type="primary", use_container_width=True)
    _run_uploaded_case(uploaded_files, run_clicked)
    render_upload_feedback()
    render_upload_landing()


def _request_records() -> list:
    requests = st.session_state.get("vendor_requests", {})
    order = st.session_state.get("vendor_request_order", [])
    return [requests[request_id] for request_id in order if request_id in requests]


def _request_record(request_id: str):
    return st.session_state.get("vendor_requests", {}).get(request_id)


def _add_uploaded_request(packet, uploaded_case, upload_workspace: Path) -> str:
    counter = int(st.session_state.get("upload_request_counter", 0)) + 1
    st.session_state["upload_request_counter"] = counter
    request_id = "uploaded_%03d" % counter
    display_number = int(st.session_state.get("next_request_display_number", len(CASE_OPTIONS) + 1))
    st.session_state["next_request_display_number"] = display_number + 1
    display_id = "VR-%03d" % display_number
    record = {
        "request_id": request_id,
        "display_id": display_id,
        "packet": packet,
        "source": "Uploaded",
        "submitted_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "uploaded_case": uploaded_case,
        "workspace": str(upload_workspace),
    }
    st.session_state.setdefault("vendor_requests", {})[request_id] = record
    st.session_state.setdefault("vendor_request_order", []).append(request_id)
    return request_id


def _delete_request(request_id: str) -> None:
    record = st.session_state.get("vendor_requests", {}).pop(request_id, None)
    st.session_state["vendor_request_order"] = [
        item for item in st.session_state.get("vendor_request_order", []) if item != request_id
    ]
    if record and record.get("workspace"):
        workspace = Path(record["workspace"])
        if workspace.exists():
            shutil.rmtree(workspace, ignore_errors=True)


def _clear_uploaded_workspaces() -> None:
    for record in _request_records():
        workspace = record.get("workspace")
        if workspace and Path(workspace).exists():
            shutil.rmtree(workspace, ignore_errors=True)


def _run_uploaded_case(uploaded_files, run_clicked: bool) -> None:
    if not run_clicked:
        return
    if not uploaded_files:
        st.error("Upload the required vendor package files before running triage.")
        return

    upload_workspace = Path(tempfile.mkdtemp(prefix="accelerant_vendor_upload_"))
    st.session_state["upload_workspace"] = str(upload_workspace)

    with st.status("Preparing uploaded package", expanded=False) as status:
        try:
            uploaded_case = stage_uploaded_case(
                _uploaded_artifacts(uploaded_files),
                upload_workspace,
                template_package_root=PACKAGE_ROOT,
            )
        except Exception as exc:
            st.session_state["packet"] = None
            st.session_state["packet_context"] = "upload:error"
            st.session_state["input_mode"] = "Submit New Request"
            st.session_state["upload_feedback"] = {
                "error": "New package could not be prepared: %s" % exc,
                "warnings": [],
            }
            status.update(label="New package could not be prepared", state="error")
            return
        if not uploaded_case.is_ready:
            st.session_state["packet"] = None
            st.session_state["uploaded_case"] = uploaded_case
            st.session_state["packet_context"] = "upload:incomplete"
            st.session_state["input_mode"] = "Submit New Request"
            errors = list(uploaded_case.blocking_errors)
            if uploaded_case.missing_roles:
                errors.append(
                    "Missing required files: %s."
                    % ", ".join(missing_role_labels(uploaded_case.missing_roles))
                )
            st.session_state["upload_feedback"] = {
                "error": " ".join(errors) or "New package is not complete enough for triage.",
                "warnings": (
                    ["Unmatched files: %s." % ", ".join(uploaded_case.unmatched_files)]
                    if uploaded_case.unmatched_files
                    else []
                )
                + uploaded_case.warnings,
            }
            status.update(label="New package incomplete", state="error")
            return
        status.update(label="Running triage", state="running")
        packet = run_case(uploaded_case.case_dir)
        request_id = _add_uploaded_request(packet, uploaded_case, upload_workspace)
        st.session_state["active_view"] = "detail"
        st.session_state["active_request_id"] = request_id
        st.session_state["upload_feedback"] = None
        status.update(label="Triage complete", state="complete")
        st.rerun()


def render_packet(packet, request_id: str = None) -> None:
    status_label = _status_label(packet)
    if packet.status == "blocked":
        st.error("%s - %s" % (status_label, packet.status_reason))
    elif packet.status == "review_required":
        st.warning("%s - %s" % (status_label, packet.status_reason))
    else:
        st.success("%s - %s" % (status_label, packet.status_reason))

    st.caption("Vendor")
    st.subheader(packet.facts.vendor_name)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("ACV", _money(packet.facts.annual_contract_value))
    col2.metric("TCV", _money(packet.facts.total_contract_value.total_contract_value))
    col3.metric("Budget", packet.facts.budget.status.title())
    col4.metric("Risk", packet.facts.risk.tier.title())

    render_action_cockpit(packet)
    with st.expander("Drafts", expanded=False):
        render_ai_assisted_drafts(packet, key_suffix=request_id or packet.case_id)
    render_audit_details(packet)


def render_action_cockpit(packet) -> None:
    rows = _review_action_rows(packet)
    next_action = rows[0]["Action"] if rows else "Route decision packet to required reviewers"
    next_owner = rows[0]["Owner"] if rows else _next_owner(packet)

    st.subheader("Decision")
    decision_col, action_col, owner_col = st.columns([0.8, 1.6, 1])
    decision_col.metric("Decision", _status_label(packet))
    action_col.markdown("**Next action**")
    action_col.write(next_action)
    owner_col.markdown("**Owner**")
    owner_col.write(next_owner)

    render_reviewer_brief(packet, show_validation=False)

    follow_up_col, route_col = st.columns([1.25, 1])
    with follow_up_col:
        render_required_follow_up(packet)
    with route_col:
        render_human_route(packet, show_guardrails=False)


def render_audit_details(packet) -> None:
    with st.expander("Audit details", expanded=False):
        context_tab, findings_tab, evidence_tab, workflow_tab, trace_tab, exports_tab = st.tabs(
            ["Context", "Findings", "Evidence", "Workflow", "Trace", "Exports"]
        )
        with context_tab:
            render_context(packet)
        with findings_tab:
            render_findings(packet, allow_expanders=False)
        with evidence_tab:
            render_evidence(packet)
        with workflow_tab:
            render_workflow_progress(packet, allow_expanders=False)
        with trace_tab:
            render_trace(packet, allow_expanders=False)
        with exports_tab:
            render_exports(packet)


def render_workflow_progress(packet, allow_expanders: bool = True) -> None:
    trace_tools = {entry.tool_name for entry in packet.trace}
    rows = [
        _workflow_row(
            "Parse and extract inputs",
            [
                "parse_intake_workbook",
                "parse_vendor_email",
                "parse_quote_csv",
                "parse_security_questionnaire",
                "parse_contract_pdf",
            ],
            trace_tools,
            "Required source files parsed into evidence-backed facts.",
        ),
        _workflow_row(
            "Validate package",
            ["parse_case_inventory", "detect_missing_information"],
            trace_tools,
            _package_validation_note(packet),
        ),
        _workflow_row(
            "Normalize case facts",
            ["classify_data_sensitivity"],
            trace_tools,
            "%s, %s ACV, %s risk."
            % (packet.facts.vendor_name, _money(packet.facts.annual_contract_value), packet.facts.risk.tier),
        ),
        _workflow_row(
            "Run deterministic helper tools",
            ["lookup_budget", "check_existing_vendor", "calculate_total_contract_value"],
            trace_tools,
            "Budget, duplicate vendor, and total contract value checks completed.",
        ),
        _workflow_row(
            "Determine approvals and risk tier",
            ["run_policy_checks", "determine_required_approvals"],
            trace_tools,
            "%s reviewer route." % len(packet.approval_route.required_reviewers),
        ),
        _workflow_row(
            "Prepare outputs",
            ["draft_human_review_messages", "prepare_reviewer_synthesis"],
            trace_tools,
            "Decision packet, reviewer brief, drafts, trace, and workbook exports are available.",
        ),
        {
            "Stage": "Human approval gate",
            "Status": "Required",
            "Function calls": "human_review",
            "Result": "Procurement owner reviews, edits, approves, or rejects.",
        },
    ]
    workflow_df = pd.DataFrame(rows)
    st.subheader("Triage Workflow")
    st.dataframe(
        workflow_df[["Stage", "Status", "Result"]],
        use_container_width=True,
        hide_index=True,
    )
    if allow_expanders:
        with st.expander("Function calls captured in trace"):
            st.dataframe(
                workflow_df[["Stage", "Function calls"]],
                use_container_width=True,
                hide_index=True,
            )
    else:
        st.subheader("Function calls captured in trace")
        st.dataframe(
            workflow_df[["Stage", "Function calls"]],
            use_container_width=True,
            hide_index=True,
        )


def render_required_follow_up(packet) -> None:
    st.subheader("Required Vendor Follow-up")
    rows = _review_action_rows(packet)
    if not rows:
        st.success("No missing information or blocking follow-up detected.")
        return
    display_rows = [
        {
            "Request": row["Action"],
            "Owner": row["Owner"],
            "Why needed": row["Why"],
            "Evidence": row["Evidence"] or "n/a",
        }
        for row in rows[:5]
    ]
    st.dataframe(pd.DataFrame(display_rows), use_container_width=True, hide_index=True)
    if len(rows) > len(display_rows):
        st.caption("%s additional follow-up item(s) included in Audit details." % (len(rows) - len(display_rows)))


def render_human_route(packet, show_guardrails: bool = True) -> None:
    st.subheader("Internal Review Route")
    st.caption("Routing recommendation only. No approval, spend commitment, or external send has occurred.")
    if packet.approval_route.required_reviewers:
        st.write(" > ".join(packet.approval_route.required_reviewers))
    else:
        st.write("Procurement owner")
    if show_guardrails:
        with st.expander("Guardrails enforced"):
            for action in packet.approval_route.prohibited_actions:
                st.write("- %s" % action)


def render_exports(packet) -> None:
    cols = st.columns(4)
    cols[0].download_button(
        "JSON packet",
        data=json.dumps(packet.model_dump(mode="json"), indent=2, sort_keys=True),
        file_name="%s_decision_packet.json" % packet.case_id,
        mime="application/json",
        use_container_width=True,
    )
    cols[1].download_button(
        "Trace",
        data=json.dumps([entry.model_dump(mode="json") for entry in packet.trace], indent=2, sort_keys=True),
        file_name="%s_trace.json" % packet.case_id,
        mime="application/json",
        use_container_width=True,
    )
    cols[2].download_button(
        "Markdown brief",
        data=_markdown_brief(packet),
        file_name="%s_brief.md" % packet.case_id,
        mime="text/markdown",
        use_container_width=True,
    )
    cols[3].download_button(
        "Workbook",
        data=_triage_workbook_bytes(packet),
        file_name="%s_triage_workbook.xlsx" % packet.case_id,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def render_upload_details(uploaded_case, expanded: bool = False) -> None:
    if not uploaded_case:
        return
    with st.expander("Staged package mapping", expanded=expanded):
        role_matches = getattr(uploaded_case, "role_matches", [])
        optional_matches = getattr(uploaded_case, "optional_matches", [])
        warnings = getattr(uploaded_case, "warnings", [])
        unmatched_files = getattr(uploaded_case, "unmatched_files", [])
        if role_matches:
            rows = [
                {
                    "Role": match.role.replace("_", " ").title(),
                    "Uploaded file": match.uploaded_name,
                    "Staged file": match.staged_name,
                }
                for match in role_matches
            ]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        if optional_matches:
            st.caption("Support artifacts")
            support_rows = [
                {
                    "Artifact": _artifact_label(match.role),
                    "Uploaded file": match.uploaded_name,
                    "Staged file": match.staged_name,
                }
                for match in optional_matches
            ]
            st.dataframe(pd.DataFrame(support_rows), use_container_width=True, hide_index=True)
        if warnings:
            for warning in warnings:
                st.warning(warning)
        if unmatched_files:
            st.caption("Ignored files: %s" % ", ".join(unmatched_files))


def render_package_delta(packet, uploaded_case, show_title: bool = True) -> None:
    if not uploaded_case:
        return
    if show_title:
        st.subheader("Package Delta")
    st.caption(
        "Uploaded files created this request record. They do not change seeded requests."
    )

    baseline_id, baseline_packet = _matching_sample_baseline(packet)
    required_count = len(getattr(uploaded_case, "role_matches", []))
    support_count = len(getattr(uploaded_case, "optional_matches", []))
    remaining_blockers = len([finding for finding in packet.findings if finding.severity == "blocker"])

    cols = st.columns(4)
    cols[0].metric("Required files", required_count)
    cols[1].metric("Support artifacts", support_count)
    cols[2].metric("Remaining requests", len(packet.missing_information))
    cols[3].metric("Remaining blockers", remaining_blockers)

    if baseline_packet:
        resolved = _resolved_missing_items(baseline_packet, packet)
        st.write("Matched baseline: %s - %s." % (baseline_id, baseline_packet.facts.vendor_name))
        if resolved:
            st.success("Resolved since baseline: %s." % "; ".join(resolved))
        else:
            st.info("No baseline missing-information items were resolved by this upload.")
    else:
        st.info("No matching sample baseline found; treating this as a net-new vendor package.")

    required_rows = [
        {
            "Role": match.role.replace("_", " ").title(),
            "Uploaded file": match.uploaded_name,
            "Staged as": match.staged_name,
        }
        for match in getattr(uploaded_case, "role_matches", [])
    ]
    support_rows = [
        {
            "Artifact": _artifact_label(match.role),
            "Uploaded file": match.uploaded_name,
            "Staged as": match.staged_name,
        }
        for match in getattr(uploaded_case, "optional_matches", [])
    ]

    delta_tab, files_tab = st.tabs(["Review delta", "Uploaded files"])
    with delta_tab:
        if packet.missing_information:
            st.markdown("**Still needed**")
            st.dataframe(_missing_rows(packet), use_container_width=True, hide_index=True)
        else:
            st.success("No missing information remains.")
        blocker_rows = _blocker_rows(packet)
        if blocker_rows:
            st.markdown("**Remaining blockers**")
            st.dataframe(pd.DataFrame(blocker_rows), use_container_width=True, hide_index=True)
        else:
            st.success("No blocking findings remain.")
    with files_tab:
        if required_rows:
            st.markdown("**Required files used for this temporary case**")
            st.dataframe(pd.DataFrame(required_rows), use_container_width=True, hide_index=True)
        if support_rows:
            st.markdown("**Support artifacts recognized**")
            st.dataframe(pd.DataFrame(support_rows), use_container_width=True, hide_index=True)
        if not required_rows and not support_rows:
            st.write("No uploaded files were staged.")


def render_upload_feedback() -> None:
    feedback = st.session_state.get("upload_feedback")
    if not feedback:
        return
    st.error(feedback["error"])
    for warning in feedback.get("warnings", []):
        st.warning(warning)


def render_upload_landing() -> None:
    expected = pd.DataFrame(
        [
            {"Role": "Intake", "Required": True, "Accepted": ".xlsx"},
            {"Role": "Quote or order form", "Required": True, "Accepted": ".csv"},
            {"Role": "Contract excerpt", "Required": True, "Accepted": ".pdf"},
            {"Role": "Security questionnaire", "Required": True, "Accepted": ".md"},
            {"Role": "Vendor email", "Required": True, "Accepted": ".txt"},
            {"Role": "Support artifacts", "Required": False, "Accepted": ".pdf, .md, .txt"},
        ]
    )
    st.dataframe(expected, use_container_width=True, hide_index=True)


def render_context(packet) -> None:
    left, right = st.columns(2)
    with left:
        st.subheader("Commercial")
        commercial = {
            "Requesting team": packet.facts.requesting_team,
            "Business owner": packet.facts.business_owner,
            "Cost center": packet.facts.cost_center,
            "Request type": packet.facts.renewal_or_new_vendor,
            "Term": "%s months" % packet.facts.contract_term_months,
            "Payment terms": packet.facts.payment_terms,
            "One-time fees": _money(packet.facts.quote.one_time_fees),
        }
        st.table(pd.DataFrame(commercial.items(), columns=["Field", "Value"]))

    with right:
        st.subheader("Risk Drivers")
        if packet.facts.risk.reasons:
            for reason in packet.facts.risk.reasons:
                st.write("- %s" % reason)
        else:
            st.write("No elevated risk drivers detected.")

    st.subheader("Missing Information")
    if packet.missing_information:
        st.dataframe(_missing_rows(packet), use_container_width=True, hide_index=True)
    else:
        st.write("No missing information detected.")


def render_overview(packet) -> None:
    render_reviewer_brief(packet)
    render_context(packet)


def render_reviewer_brief(packet, show_validation: bool = True) -> None:
    st.subheader("Reviewer Brief")
    synthesis = getattr(packet, "synthesis", None)
    if not synthesis:
        st.markdown(_summary_html(packet.summary), unsafe_allow_html=True)
        return
    st.markdown(_summary_html(synthesis.executive_summary), unsafe_allow_html=True)
    st.caption(
        "Built from the validated decision packet. Policy status, risk, budget, and routing remain deterministic."
    )
    if show_validation:
        st.caption("Synthesis source: %s" % synthesis.model_name)
        with st.expander("Synthesis validation"):
            st.write("Status: %s" % synthesis.validation_status)
            st.write("Source: structured decision packet only")
            st.write("Evidence cited: %s" % (", ".join(synthesis.cited_evidence_ids) or "n/a"))
            if synthesis.validation_errors:
                for error in synthesis.validation_errors:
                    st.warning(error)


def render_ai_assisted_drafts(packet, key_suffix: str = None) -> None:
    st.subheader("Drafts")
    synthesis = getattr(packet, "synthesis", None)
    use_synthesis = synthesis and synthesis.validation_status.startswith("passed")
    vendor_body = (
        synthesis.vendor_follow_up_draft
        if use_synthesis
        else _draft_body(packet, "vendor")
    )
    internal_body = (
        synthesis.internal_note_draft
        if use_synthesis
        else _draft_body(packet, "internal")
    )
    source = synthesis.model_name if synthesis else "deterministic packet draft"
    st.caption("Draft source: %s. Human approval is required before external use." % source)
    key_suffix = key_suffix or packet.case_id

    vendor_tab, internal_tab = st.tabs(["Vendor follow-up", "Internal note"])
    with vendor_tab:
        st.text_area(
            "Draft vendor follow-up",
            value=vendor_body,
            height=220,
            key="ai_vendor_follow_up_%s" % key_suffix,
        )
    with internal_tab:
        st.text_area(
            "Draft internal note",
            value=internal_body,
            height=220,
            key="ai_internal_note_%s" % key_suffix,
        )


def render_findings(packet, allow_expanders: bool = True) -> None:
    rows = [
        {
            "Function": finding.function,
            "Severity": finding.severity,
            "Trigger": finding.trigger,
            "Owner": finding.required_owner,
            "Recommended action": finding.recommended_action,
            "Evidence": len(finding.evidence_ids),
        }
        for finding in packet.findings
    ]
    findings_df = pd.DataFrame(rows)
    blocker_rows = findings_df[findings_df["Severity"] == "blocker"] if not findings_df.empty else findings_df
    if not blocker_rows.empty:
        st.subheader("Blocking Issues")
        st.dataframe(
            blocker_rows[["Function", "Trigger", "Owner", "Recommended action", "Evidence"]],
            use_container_width=True,
            hide_index=True,
        )

    st.subheader("All Findings")
    st.dataframe(findings_df, use_container_width=True, hide_index=True)

    if allow_expanders:
        with st.expander("Detailed policy rationale"):
            render_policy_rationale(packet)
    else:
        st.subheader("Detailed policy rationale")
        render_policy_rationale(packet)


def render_policy_rationale(packet) -> None:
    for function in sorted({finding.function for finding in packet.findings}):
        st.markdown("**%s**" % function)
        for finding in [item for item in packet.findings if item.function == function]:
            st.markdown("- **%s**" % finding.trigger)
            st.write(finding.why_it_matters)
            st.caption("Owner: %s | Severity: %s" % (finding.required_owner, finding.severity))
            st.caption("Policy: %s" % "; ".join(finding.policy_refs))


def render_evidence(packet) -> None:
    rows = [
        {
            "ID": item.id,
            "Source type": item.source_type,
            "Source file": Path(item.source_file).name,
            "Location": item.location,
            "Parsed value": item.parsed_value or "",
            "Snippet": item.snippet,
        }
        for item in packet.evidence
    ]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def render_drafts(packet) -> None:
    acknowledged = st.checkbox(
        "I understand these are drafts and require human approval before use.",
        value=False,
    )
    for draft in packet.drafts:
        st.subheader("%s draft" % draft.audience.title())
        st.text_input(
            "Subject",
            value=draft.subject,
            key="subject_%s_%s" % (packet.case_id, draft.audience),
            disabled=not acknowledged,
        )
        st.text_area(
            "Body",
            value=draft.body,
            height=220,
            key="body_%s_%s" % (packet.case_id, draft.audience),
            disabled=not acknowledged,
        )


def render_trace(packet, allow_expanders: bool = True) -> None:
    rows = [
        {
            "Step": index,
            "Tool": entry.tool_name,
            "Status": entry.status,
            "Duration ms": entry.duration_ms,
            "Requirements": ", ".join(entry.requirement_ids),
            "Evidence IDs": ", ".join(entry.evidence_ids),
        }
        for index, entry in enumerate(packet.trace, start=1)
    ]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    if allow_expanders:
        with st.expander("Raw trace JSON"):
            st.json([entry.model_dump(mode="json") for entry in packet.trace])
    else:
        st.subheader("Raw trace JSON")
        st.json([entry.model_dump(mode="json") for entry in packet.trace])


def _matching_sample_baseline(packet):
    uploaded_vendor = _normalize_vendor_name(packet.facts.vendor_name)
    for case_id in CASE_OPTIONS:
        baseline = _run_sample_case_cached(case_id, _synthesis_runtime_signature())
        if _normalize_vendor_name(baseline.facts.vendor_name) == uploaded_vendor:
            return case_id, baseline
    return None, None


def _resolved_missing_items(baseline_packet, uploaded_packet) -> list:
    baseline_missing = {item.item for item in baseline_packet.missing_information}
    uploaded_missing = {item.item for item in uploaded_packet.missing_information}
    return sorted(baseline_missing - uploaded_missing)


def _blocker_rows(packet) -> list:
    return [
        {
            "Function": finding.function,
            "Issue": finding.trigger,
            "Owner": finding.required_owner,
            "Action": finding.recommended_action,
        }
        for finding in packet.findings
        if finding.severity == "blocker"
    ]


def _next_owner(packet) -> str:
    if packet.missing_information:
        return packet.missing_information[0].owner
    if packet.approval_route.required_reviewers:
        return packet.approval_route.required_reviewers[-1]
    return "Procurement owner"


def _workflow_row(stage: str, tools: list, trace_tools: set, outcome: str) -> dict:
    completed = [tool for tool in tools if tool in trace_tools]
    status = "Complete" if len(completed) == len(tools) else "Pending"
    return {
        "Stage": stage,
        "Status": status,
        "Function calls": ", ".join(tools),
        "Result": outcome,
    }


def _package_validation_note(packet) -> str:
    if packet.missing_information:
        return "%s open request(s); not complete enough for approval readiness." % len(packet.missing_information)
    return "Complete enough for triage."


def _artifact_label(key: str) -> str:
    labels = {
        "soc2_type2": "SOC 2 Type II",
        "data_processing_agreement": "Data Processing Agreement",
        "subprocessor_list": "Subprocessor list",
        "tax_form": "Tax form",
        "vendor_setup_form": "Vendor setup form",
        "ai_training_opt_out": "AI training opt-out",
        "incident_response_summary": "Incident response summary",
        "statement_of_work": "Statement of work",
    }
    return labels.get(key, key.replace("_", " ").title())


def _triage_workbook_bytes(packet) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _append_rows(
        summary,
        [
            ("Case ID", packet.case_id),
            ("Vendor", packet.facts.vendor_name),
            ("Status", packet.status),
            ("Status reason", packet.status_reason),
            ("Risk", packet.facts.risk.tier),
            ("ACV", packet.facts.annual_contract_value),
            ("TCV", packet.facts.total_contract_value.total_contract_value),
            ("Budget", packet.facts.budget.status),
            ("Structured summary", packet.summary),
            ("Reviewer brief", _reviewer_summary(packet)),
        ],
    )

    missing = wb.create_sheet("Missing Info")
    missing.append(["Item", "Owner", "Why needed", "Evidence IDs"])
    for item in packet.missing_information:
        missing.append([item.item, item.owner, item.why_needed, ", ".join(item.evidence_ids)])

    findings = wb.create_sheet("Findings")
    findings.append(["ID", "Function", "Severity", "Trigger", "Owner", "Action", "Evidence IDs", "Policy refs"])
    for finding in packet.findings:
        findings.append(
            [
                finding.id,
                finding.function,
                finding.severity,
                finding.trigger,
                finding.required_owner,
                finding.recommended_action,
                ", ".join(finding.evidence_ids),
                "; ".join(finding.policy_refs),
            ]
        )

    route = wb.create_sheet("Approval Route")
    route.append(["Order", "Reviewer"])
    for index, reviewer in enumerate(packet.approval_route.required_reviewers, start=1):
        route.append([index, reviewer])
    route.append([])
    route.append(["Prohibited action"])
    for action in packet.approval_route.prohibited_actions:
        route.append([action])

    trace = wb.create_sheet("Trace")
    trace.append(["Step", "Tool", "Status", "Duration ms", "Requirements", "Evidence IDs"])
    for index, entry in enumerate(packet.trace, start=1):
        trace.append(
            [
                index,
                entry.tool_name,
                entry.status,
                entry.duration_ms,
                ", ".join(entry.requirement_ids),
                ", ".join(entry.evidence_ids),
            ]
        )

    if getattr(packet, "synthesis", None):
        synthesis = wb.create_sheet("Synthesis")
        _append_rows(
            synthesis,
            [
                ("Mode", packet.synthesis.synthesis_mode),
                ("Model", packet.synthesis.model_name),
                ("Validation", packet.synthesis.validation_status),
                ("Executive summary", packet.synthesis.executive_summary),
                ("Vendor follow-up draft", packet.synthesis.vendor_follow_up_draft),
                ("Internal note draft", packet.synthesis.internal_note_draft),
                ("Evidence IDs", ", ".join(packet.synthesis.cited_evidence_ids)),
                ("Validation errors", "; ".join(packet.synthesis.validation_errors)),
            ],
        )

    for ws in wb.worksheets:
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            ws.column_dimensions[column_cells[0].column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _append_rows(sheet, rows) -> None:
    for label, value in rows:
        sheet.append([label, value])


def _status_label(packet) -> str:
    if packet.status == "blocked" and packet.facts.risk.tier == "low":
        return "LOW-RISK SETUP DOCS MISSING"
    if packet.status == "blocked":
        return "BLOCKED"
    if packet.status == "review_required":
        return "READY FOR REVIEW"
    return "READY"


def _primary_next_action(packet) -> str:
    rows = _review_action_rows(packet)
    if rows:
        return rows[0]["Action"]
    return "Route decision packet to required reviewers"


def _review_action_rows(packet) -> list:
    rows = []
    seen = set()

    for item in packet.missing_information:
        key = ("missing", item.item, item.owner)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "Action": item.item,
                "Owner": item.owner,
                "Why": item.why_needed,
                "Evidence": ", ".join(item.evidence_ids),
            }
        )

    for finding in packet.findings:
        if finding.severity not in {"blocker", "review_required"}:
            continue
        if packet.missing_information and finding.trigger == "Required onboarding materials are missing.":
            continue
        key = ("finding", finding.recommended_action, finding.required_owner)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "Action": finding.recommended_action,
                "Owner": finding.required_owner,
                "Why": finding.trigger,
                "Evidence": ", ".join(finding.evidence_ids),
            }
        )

    return rows[:8]


def _missing_rows(packet) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Item": item.item,
                "Owner": item.owner,
                "Why needed": item.why_needed,
                "Evidence IDs": ", ".join(item.evidence_ids),
            }
            for item in packet.missing_information
        ]
    )


def _money(value: float) -> str:
    return "$%s" % format(value, ",.0f")


def _normalize_vendor_name(value: str) -> str:
    return (value or "").lower().replace(",", "").replace(".", "").strip()


def _markdown_brief(packet) -> str:
    missing = "\n".join("- %s (%s)" % (item.item, item.owner) for item in packet.missing_information)
    findings = "\n".join(
        "- [%s] %s: %s" % (finding.severity, finding.function, finding.trigger)
        for finding in packet.findings
    )
    route = "\n".join(
        "%s. %s" % (index, reviewer)
        for index, reviewer in enumerate(packet.approval_route.required_reviewers, start=1)
    )
    return (
        "# %s Vendor Triage Brief\n\n"
        "Status: %s\n\n"
        "Risk: %s\n\n"
        "ACV: %s\n\n"
        "TCV: %s\n\n"
        "## Reviewer Brief\n\n%s\n\n"
        "## Structured Summary\n\n%s\n\n"
        "## Missing Information\n\n%s\n\n"
        "## Required Human Route\n\n%s\n\n"
        "## Findings\n\n%s\n"
    ) % (
        packet.facts.vendor_name,
        packet.status,
        packet.facts.risk.tier,
        _money(packet.facts.annual_contract_value),
        _money(packet.facts.total_contract_value.total_contract_value),
        _reviewer_summary(packet),
        packet.summary,
        missing or "None",
        route,
        findings or "None",
    )


def _reviewer_summary(packet) -> str:
    synthesis = getattr(packet, "synthesis", None)
    if synthesis and synthesis.validation_status.startswith("passed"):
        return synthesis.executive_summary
    return packet.summary


def _draft_body(packet, audience: str) -> str:
    for draft in packet.drafts:
        if draft.audience == audience:
            return draft.body
    return ""


def _summary_html(value: str) -> str:
    escaped = html.escape(value).replace("$", "&#36;")
    return '<p class="summary-text">%s</p>' % escaped


def _uploaded_artifacts(uploaded_files) -> list:
    return [
        UploadedArtifact(name=uploaded_file.name, content=uploaded_file.getvalue())
        for uploaded_file in uploaded_files
    ]


def _synthesis_runtime_signature() -> str:
    provider = os.getenv("OPENAI_SYNTHESIS_PROVIDER", "deterministic").lower()
    model = os.getenv("OPENAI_SYNTHESIS_MODEL", "")
    key_state = "key" if os.getenv("OPENAI_API_KEY") else "no-key"
    return "%s|%s|%s" % (provider, model, key_state)


st.markdown(
    """
    <style>
    .summary-text {
        font-size: 1rem;
        line-height: 1.55;
        margin-bottom: 1.25rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


if __name__ == "__main__":
    main()
